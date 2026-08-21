"""
fill_bingo_cards.py

Reads Formspree submissions (exported to Excel) and the bilingual prompts.json
source file, then fills each participant's 5x5 bingo card in bingo_forms.xlsm
- prompt text goes into the grid in the participant's chosen language
- the 24 chosen prompts are shuffled into the 24 non-center cells
- the center cell (D4) is left untouched / blank (the free space)
- each cell also gets the prompt's short description attached as a cell
  comment (hover note), so it's there for reference but doesn't clutter the
  card itself
- each card is also saved as a (5,5) numpy array to forms/<name>.npy, holding
  the prompt IDs in the same grid layout (center cell is an empty string)

USAGE
-----
    python fill_bingo_cards.py

By default this expects, in the same folder:
    raw_submissions.xlsx   (columns: name | language | selected_ids)
    prompts.json           (the same source file used by the website)
    bingo_forms.xlsm        (existing workbook, one sheet per participant)

...and writes a new file called bingo_forms_filled.xlsm, so your original
template is never overwritten by accident. Pass --in-place if you want it to
save back into bingo_forms.xlsm directly instead.

All paths can be overridden, e.g.:
    python fill_bingo_cards.py --submissions raw_submissions.xlsx \
        --prompts prompts.json --bingo bingo_forms.xlsm --output filled.xlsm

Requires: pip install openpyxl numpy
"""

import argparse
import json
import random
import re
import sys
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.comments import Comment

# The 5x5 grid lives in B2:F6, with D4 as the fixed free/center cell.
GRID_CELLS = [f"{col}{row}" for row in range(2, 7) for col in "BCDEF"]
CENTER_CELL = "D4"
NON_CENTER_CELLS = [c for c in GRID_CELLS if c != CENTER_CELL]


def load_prompts(prompts_path):
    """Load prompts.json into a dict keyed by prompt id."""
    with open(prompts_path, encoding="utf-8") as f:
        data = json.load(f)
    prompts = {p["id"]: p for p in data["prompts"]}
    if not prompts:
        raise ValueError(f"No prompts found in {prompts_path}")
    return prompts


def load_submissions(submissions_path):
    """Read raw_submissions.xlsx: column A name, B language, C selected ids."""
    wb = openpyxl.load_workbook(submissions_path, data_only=True)
    ws = wb.active

    submissions = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name, language, ids_raw = (row + (None, None, None))[:3]
        if name is None or str(name).strip() == "":
            continue
        if ids_raw is None:
            print(f"  WARNING: row {row_num} ({name}) has no selected prompts, skipping")
            continue

        ids = [x.strip().zfill(2) for x in str(ids_raw).split(",") if x.strip()]
        language = str(language).strip().lower() if language else "en"
        if language not in ("en", "de"):
            print(f"  WARNING: row {row_num} ({name}) has unknown language '{language}', defaulting to 'en'")
            language = "en"

        submissions.append({"row": row_num, "name": str(name).strip(), "language": language, "ids": ids})
    return submissions


def fill_bingo_card(ws, ids, prompts, language):
    """Write shuffled prompt titles into ws's grid; center stays blank.

    Returns a (5, 5) numpy array of prompt IDs in the same grid layout
    (row 0 = row 2 of the sheet, col 0 = column B), with the center cell
    as an empty string.
    """
    if len(ids) != len(NON_CENTER_CELLS):
        raise ValueError(
            f"expected {len(NON_CENTER_CELLS)} selected prompts, got {len(ids)}"
        )

    missing = [pid for pid in ids if pid not in prompts]
    if missing:
        raise KeyError(f"prompt id(s) not found in prompts.json: {', '.join(missing)}")

    shuffled_ids = ids[:]
    random.shuffle(shuffled_ids)

    cell_to_id = {CENTER_CELL: 0}
    for cell_ref, prompt_id in zip(NON_CENTER_CELLS, shuffled_ids):
        prompt = prompts[prompt_id]
        text = prompt.get(language, prompt["en"])
        cell = ws[cell_ref]
        cell.value = text["title"]
        cell.comment = Comment(text["description"], "Bingo Generator")
        cell_to_id[cell_ref] = prompt_id

    grid = np.empty((5, 5), dtype=int)
    for row_idx, row in enumerate(range(2, 7)):
        for col_idx, col in enumerate("BCDEF"):
            grid[row_idx, col_idx] = int(cell_to_id[f"{col}{row}"])
    return grid


def sanitize_filename(name):
    """Make a participant name safe to use as a filename."""
    safe = re.sub(r"[^\w\-. ]", "_", name).strip()
    return safe or "unnamed"


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--submissions", default="raw_submissions.xlsx", help="Formspree export (xlsx)")
    parser.add_argument("--prompts", default="prompts.json", help="Bilingual prompt source file")
    parser.add_argument("--bingo", default="bingo_forms.xlsm", help="Existing workbook with empty bingo sheets")
    parser.add_argument("--output", default=None, help="Where to save the filled workbook")
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Save back into --bingo directly instead of a new file",
    )
    parser.add_argument(
        "--npy-dir",
        default="forms",
        help="Folder to save each participant's (5,5) numpy array into",
    )
    args = parser.parse_args()

    if args.output:
        output_path = args.output
    elif args.in_place:
        output_path = args.bingo
    else:
        bingo_path = Path(args.bingo)
        output_path = str(bingo_path.with_name(bingo_path.stem + "_filled" + bingo_path.suffix))

    print(f"Loading prompts from {args.prompts} ...")
    prompts = load_prompts(args.prompts)
    print(f"  {len(prompts)} prompts loaded")

    print(f"Loading submissions from {args.submissions} ...")
    submissions = load_submissions(args.submissions)
    print(f"  {len(submissions)} submissions loaded")

    print(f"Opening {args.bingo} (keeping macros) ...")
    wb = openpyxl.load_workbook(args.bingo, keep_vba=True)

    npy_dir = Path(args.npy_dir)
    npy_dir.mkdir(parents=True, exist_ok=True)

    filled, skipped = 0, []
    for sub in submissions:
        if sub["name"] not in wb.sheetnames:
            print(f"  ERROR: no sheet named '{sub['name']}' found — skipping (row {sub['row']})")
            skipped.append(sub["name"])
            continue

        ws = wb[sub["name"]]
        try:
            grid = fill_bingo_card(ws, sub["ids"], prompts, sub["language"])
            npy_path = npy_dir / f"{sanitize_filename(sub['name'])}.npy"
            np.save(npy_path, grid)
            print(f"  OK: {sub['name']} ({sub['language']}, {len(sub['ids'])} prompts) -> {npy_path}")
            filled += 1
        except (ValueError, KeyError) as e:
            print(f"  ERROR filling '{sub['name']}': {e} — skipping")
            skipped.append(sub["name"])

    wb.save(output_path)

    print()
    print(f"Done: {filled} card(s) filled, {len(skipped)} skipped.")
    if skipped:
        print("Skipped:", ", ".join(skipped))
    print(f"Saved to {output_path}")

    if not output_path.lower().endswith(".xlsm"):
        print("NOTE: output filename doesn't end in .xlsm — rename it back to .xlsm to keep your macros working.")


if __name__ == "__main__":
    sys.exit(main())